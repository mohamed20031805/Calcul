"""
HITS - Traitement trimestriel
==============================
Usage :
    python hits_traitement.py

1. Place tous tes fichiers .xls/.xlsx du trimestre dans le dossier 'data/'
2. Configure les 3 jours non fixes dans la section CONFIGURATION ci-dessous
3. Lance le script
4. Résultat dans 'output/HITS_rapport_<trimestre>.xlsx'
"""

import os
import sys
import glob
from datetime import date, timedelta

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ─────────────────────────────────────────────
#  CONFIGURATION  ← modifier ici avant de lancer
# ─────────────────────────────────────────────

INPUT_FOLDER  = "data"          # dossier contenant les .xls/.xlsx
OUTPUT_FOLDER = "output"        # dossier de sortie

# Trimestre traité (pour le nom du fichier de sortie)
TRIMESTRE = "Q1-2025"

# 3 jours non fixes à exclure du calcul diff_off (format YYYY-MM-DD)
JOURS_NON_FIXES = [
    # "2025-03-03",
    # "2025-06-15",
    # "2025-09-22",
]

# Noms exacts (ou partiels) des colonnes dans tes fichiers source
COL_UPDATE      = "Update Date"
COL_INTEGRATION = "Integration Date"

# Colonnes utilisées pour la Clé Unique (noms partiels acceptés)
COL_FA_NAME    = "Fund administrator name"
COL_AMOUNT     = "Amount in EUR"
COL_FA_ACCOUNT = "FA Account Number"
COL_OWNER      = "Account Owner"

# Colonne statut et valeur du statut d'ouverture
COL_STATUS         = "Status"          # nom partiel accepté
STATUS_OPEN        = "To be treated"   # statut = dossier ouvert (1 seul statut)

# Date de fin de trimestre utilisée quand une clé n'a qu'1 statut (format YYYY-MM-DD)
DATE_FIN_TRIMESTRE = "2025-03-31"

# Colonne équipe et nom de l'onglet NOK dans les fichiers source
COL_TEAM       = "Team In-Charge"   # nom partiel accepté
NOK_SHEET_NAME = "NOK"              # nom exact de l'onglet NOK dans les fichiers source

# ─────────────────────────────────────────────
#  JOURS FÉRIÉS LUXEMBOURG
# ─────────────────────────────────────────────

def _easter(year: int) -> date:
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19*a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2*e + 2*i - h - k) % 7
    m = (a + 11*h + 22*l) // 451
    month, day = divmod(h + l - 7*m + 114, 31)
    return date(year, month, day + 1)


def feriers_luxembourg(year: int) -> set:
    easter = _easter(year)
    fixed = {
        date(year, 1,  1),
        date(year, 5,  1),
        date(year, 5,  9),
        date(year, 6, 23),
        date(year, 8, 15),
        date(year, 11, 1),
        date(year, 12, 25),
        date(year, 12, 26),
    }
    mobile = {
        easter + timedelta(days=1),   # Lundi de Pâques
        easter + timedelta(days=39),  # Ascension
        easter + timedelta(days=50),  # Lundi de Pentecôte
    }
    return fixed | mobile


def build_excluded_set(extra_days: list[str]) -> set:
    excluded = set()
    for s in extra_days:
        excluded.add(date.fromisoformat(s.strip()))
    return excluded


def is_working_day(d: date, feriers_by_year: dict, extra_excluded: set) -> bool:
    if d.weekday() >= 5:
        return False
    year = d.year
    if year not in feriers_by_year:
        feriers_by_year[year] = feriers_luxembourg(year)
    if d in feriers_by_year[year]:
        return False
    if d in extra_excluded:
        return False
    return True


def diff_off(d1: date, d2: date, feriers_by_year: dict, extra_excluded: set) -> int | None:
    if pd.isna(d1) or pd.isna(d2):
        return None
    try:
        start, end = (d1, d2) if d1 <= d2 else (d2, d1)
        sign = 1 if d1 <= d2 else -1
        count = 0
        cur = start + timedelta(days=1)
        while cur <= end:
            if is_working_day(cur, feriers_by_year, extra_excluded):
                count += 1
            cur += timedelta(days=1)
        return sign * count
    except Exception:
        return None

# ─────────────────────────────────────────────
#  LECTURE & FUSION DES FICHIERS
# ─────────────────────────────────────────────

def find_col(df: pd.DataFrame, hint: str) -> str | None:
    for col in df.columns:
        if hint.lower() in str(col).lower():
            return col
    return None


def load_and_fuse(folder: str) -> pd.DataFrame:
    patterns = [
        os.path.join(folder, "*.xlsx"),
        os.path.join(folder, "*.xls"),
    ]
    files = []
    for p in patterns:
        files.extend(glob.glob(p))

    if not files:
        sys.exit(f"[ERREUR] Aucun fichier .xls/.xlsx trouvé dans '{folder}/'")

    frames = []
    for f in files:
        print(f"  Lecture : {os.path.basename(f)}")
        try:
            df = pd.read_excel(f, dtype=str)
            df["_source_file"] = os.path.basename(f)
            frames.append(df)
        except Exception as e:
            print(f"  [AVERT] Impossible de lire {f} : {e}")

    if not frames:
        sys.exit("[ERREUR] Aucun fichier lisible.")

    fused = pd.concat(frames, ignore_index=True)
    print(f"\n  {len(files)} fichier(s) fusionné(s) → {len(fused)} lignes brutes")
    return fused


def appliquer_logique_statuts(df: pd.DataFrame, col_st: str, col_ud: str, date_fin: pd.Timestamp) -> pd.DataFrame:
    """
    Pour chaque Clé Unique :
      - 1 statut  → dossier encore ouvert : Update Date remplacée par date_fin
      - 2 statuts → on garde les 2 lignes telles quelles
      - 3+ statuts → on supprime les lignes à partir du rang 3 (on garde rang 1 et 2)
    Le rang est déterminé par l'ordre croissant de Update Date au sein de la clé.
    """
    avant = len(df)

    # Trier par Clé Unique puis Update Date pour établir le rang
    df = df.sort_values(["Clé Unique", col_ud]).copy()
    df["_rang"] = df.groupby("Clé Unique").cumcount() + 1          # rang 1, 2, 3...
    df["_nb_statuts"] = df.groupby("Clé Unique")["_rang"].transform("max")

    # Logique 1 — clé avec 1 seul statut : remplacer Update Date par date fin trimestre
    masque_1 = df["_nb_statuts"] == 1
    df.loc[masque_1, col_ud] = date_fin
    nb_ouverts = masque_1.sum()

    # Logique 3 — clé avec 3+ statuts : supprimer les lignes de rang >= 3
    masque_sup = df["_rang"] >= 3
    nb_supprimes = masque_sup.sum()
    df = df[~masque_sup].copy()

    df.drop(columns=["_rang", "_nb_statuts"], inplace=True)

    print(f"  Dossiers encore ouverts (1 statut) : {nb_ouverts} → Update Date remplacée par {date_fin.date()}")
    print(f"  Lignes supprimées (rang 3+)         : {nb_supprimes}")
    print(f"  Lignes après logique statuts        : {len(df)}")
    return df


def clean_and_compute(df: pd.DataFrame) -> pd.DataFrame:
    # Trouver les colonnes dates
    col_ud = find_col(df, COL_UPDATE)
    col_ig = find_col(df, COL_INTEGRATION)

    if not col_ud:
        sys.exit(f"[ERREUR] Colonne '{COL_UPDATE}' introuvable. Colonnes dispo : {list(df.columns)}")
    if not col_ig:
        sys.exit(f"[ERREUR] Colonne '{COL_INTEGRATION}' introuvable. Colonnes dispo : {list(df.columns)}")

    # Supprimer lignes Update Date vide
    avant = len(df)
    df = df[df[col_ud].notna() & (df[col_ud].str.strip() != "")].copy()
    print(f"  Lignes supprimées (Update Date vide) : {avant - len(df)}")
    print(f"  Lignes conservées : {len(df)}")

    # Parser les dates
    df[col_ud] = pd.to_datetime(df[col_ud], errors="coerce")
    df[col_ig] = pd.to_datetime(df[col_ig], errors="coerce")

    # Clé Unique — IntegrationDate + FA Name + Amount EUR + FA Account + Owner
    col_fa   = find_col(df, COL_FA_NAME)
    col_amt  = find_col(df, COL_AMOUNT)
    col_faacc= find_col(df, COL_FA_ACCOUNT)
    col_own  = find_col(df, COL_OWNER)

    def _safe(row, col):
        return str(row[col]).strip() if col and pd.notna(row.get(col, None)) else ""

    def build_key(row):
        ig_str = row[col_ig].strftime("%Y/%m/%d %H:%M:%S") if pd.notna(row[col_ig]) else ""
        return "|".join([
            ig_str,
            _safe(row, col_fa),
            _safe(row, col_amt),
            _safe(row, col_faacc),
            _safe(row, col_own),
        ])

    df.insert(0, "Clé Unique", df.apply(build_key, axis=1))

    # Vérification doublons
    dupes = df["Clé Unique"].duplicated().sum()
    if dupes:
        print(f"  [AVERT] {dupes} clé(s) dupliquée(s) détectée(s)")
    else:
        print(f"  Clé Unique : OK — toutes les clés sont uniques")

    # ── Logique métier sur les statuts ────────────────────────────────────
    col_st = find_col(df, COL_STATUS)
    date_fin = pd.Timestamp(DATE_FIN_TRIMESTRE)

    if col_st:
        df = appliquer_logique_statuts(df, col_st, col_ud, date_fin)
    else:
        print(f"  [AVERT] Colonne statut '{COL_STATUS}' introuvable — logique statuts ignorée")

    # diff_day (calendaire) — comparaison sur la date uniquement, heure ignorée
    df["diff_day"] = (df[col_ud].dt.normalize() - df[col_ig].dt.normalize()).dt.days

    # diff_off (jours ouvrés Luxembourg)
    feriers_cache = {}
    extra_excluded = build_excluded_set(JOURS_NON_FIXES)

    df["diff_off"] = df.apply(
        lambda row: diff_off(
            row[col_ig].date() if pd.notna(row[col_ig]) else None,
            row[col_ud].date() if pd.notna(row[col_ud]) else None,
            feriers_cache,
            extra_excluded,
        ),
        axis=1,
    )

    return df, col_ud, col_ig

# ─────────────────────────────────────────────
#  STATISTIQUES
# ─────────────────────────────────────────────

def compute_stats(series: pd.Series, label: str) -> dict:
    s = series.dropna()
    if len(s) == 0:
        return {}
    return {
        "label":    label,
        "count":    len(s),
        "min":      int(s.min()),
        "max":      int(s.max()),
        "mean":     round(float(s.mean()), 2),
        "median":   float(s.median()),
        "p90":      float(s.quantile(0.90)),
        "p95":      float(s.quantile(0.95)),
        "std":      round(float(s.std()), 2),
        "neg":      int((s < 0).sum()),
        "zero":     int((s == 0).sum()),
        "pos":      int((s > 0).sum()),
    }


# ─────────────────────────────────────────────
#  STATISTIQUES PAR ÉQUIPE (Team In-Charge)
# ─────────────────────────────────────────────

def load_nok_keys(folder: str) -> set:
    """Lit tous les onglets NOK des fichiers source et retourne les Clés Uniques NOK."""
    patterns = [os.path.join(folder, "*.xlsx"), os.path.join(folder, "*.xls")]
    files = []
    for p in patterns:
        files.extend(glob.glob(p))

    nok_keys = set()
    for f in files:
        try:
            xf = pd.ExcelFile(f)
            # Chercher un onglet dont le nom contient NOK
            nok_sheet = next((s for s in xf.sheet_names if NOK_SHEET_NAME.lower() in s.lower()), None)
            if nok_sheet:
                df_nok = pd.read_excel(f, sheet_name=nok_sheet, dtype=str)
                col_key = next((c for c in df_nok.columns if "cl" in c.lower() and "unique" in c.lower()), None)
                if col_key:
                    nok_keys.update(df_nok[col_key].dropna().str.strip().tolist())
        except Exception as e:
            print(f"  [AVERT] Lecture NOK ignorée pour {os.path.basename(f)} : {e}")

    print(f"  Clés NOK chargées depuis onglets '{NOK_SHEET_NAME}' : {len(nok_keys)}")
    return nok_keys


def build_team_stats(df: pd.DataFrame, nok_keys: set) -> dict:
    """
    Préparation et calcul des KPIs par Team In-Charge.

    Étapes :
      1. Supprimer lignes Team In-Charge vide
      2. Supprimer lignes où diff_day=0 ET diff_off=0
      3. Déduplication Clé Unique → garder statut ≠ To be treated
         (si tous les statuts sont To be treated, on garde quand même la ligne)
    KPIs par groupe Team :
      A = somme diff_off
      B = nombre de lignes (après dédup)
      Moyenne = A / B
      C = nb lignes dont Clé Unique dans onglet NOK
      % = C / B × 100
    """
    col_team = find_col(df, COL_TEAM)
    col_key  = find_col(df, "Clé Unique")
    col_st   = find_col(df, COL_STATUS)

    if not col_team:
        print(f"  [AVERT] Colonne '{COL_TEAM}' introuvable — stats par équipe ignorées")
        return {}

    # Étape 1 — Supprimer lignes Team In-Charge vide
    avant = len(df)
    df = df[df[col_team].notna() & (df[col_team].astype(str).str.strip() != "")].copy()
    print(f"  Lignes supprimées (Team In-Charge vide)      : {avant - len(df)}")

    # Étape 2 — Supprimer lignes où diff_day=0 ET diff_off=0
    avant2 = len(df)
    df = df[~((df["diff_day"].fillna(0) == 0) & (df["diff_off"].fillna(0) == 0))].copy()
    print(f"  Lignes supprimées (diff_day=0 ET diff_off=0) : {avant2 - len(df)}")

    # Étape 2b — Garder uniquement les lignes où diff_off == diff_day
    avant3 = len(df)
    df = df[df["diff_off"].fillna(-1) == df["diff_day"].fillna(-2)].copy()
    print(f"  Lignes supprimées (diff_off ≠ diff_day)      : {avant3 - len(df)}")

    # Étape 3 — Déduplication Clé Unique : garder statut ≠ To be treated en priorité
    if col_st and col_key:
        df["_is_open"] = df[col_st].astype(str).str.strip() == STATUS_OPEN
        df_dedup = (
            df.sort_values("_is_open")          # False (≠ open) en premier
            .drop_duplicates(subset=[col_key], keep="first")
            .drop(columns=["_is_open"])
            .copy()
        )
        df = df.drop(columns=["_is_open"])
    else:
        df_dedup = df.drop_duplicates(subset=[col_key]).copy() if col_key else df.copy()

    print(f"  Lignes après déduplication Clé Unique        : {len(df_dedup)}")

    result = {}
    teams = sorted(df_dedup[col_team].dropna().unique())

    for team in teams:
        df_team = df_dedup[df_dedup[col_team] == team].copy()

        B       = len(df_team)
        A       = round(float(df_team["diff_off"].fillna(0).sum()), 2)
        moyenne = round(A / B, 2) if B else 0

        if col_key and nok_keys:
            C = int(df_team[col_key].isin(nok_keys).sum())
        else:
            C = 0

        pct_nok = round(C / B * 100, 2) if B else 0

        result[team] = {
            "A":         A,
            "B":         B,
            "moyenne":   moyenne,
            "C":         C,
            "pct_nok":   pct_nok,
            "df_detail": df_team,
        }

    return result


def write_team_sheet(wb, team: str, stats: dict):
    """Crée une feuille Excel pour une équipe avec ses KPIs + données détail."""
    # Nom de feuille Excel : max 31 chars, pas de caractères spéciaux
    sheet_name = str(team)[:31].replace("/","_").replace("\\","_").replace("?","").replace("*","").replace("[","").replace("]","").replace(":","")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # ── Titre équipe
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"Équipe : {team}  —  {TRIMESTRE}"
    c.font      = hdr(bold=True, color=WHITE, size=13)
    c.fill      = fill(BLUE_DARK)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    # ── Bloc KPIs
    kpis = [
        ("A",         "Somme diff_off — total jours ouvrés par équipe",  stats["A"],       "0.00"),
        ("B",         "Nombre de mouvements (lignes après dédup)",        stats["B"],       None),
        ("Moyenne",   "Moyenne Nombre de jours  (A / B)",                 stats["moyenne"], "0.00"),
        ("C",         "Nombre de mvt NON Justifiés (clés dans NOK)",      stats["C"],       None),
        ("%",         "% mvt non justifiés  (C / B × 100)",               stats["pct_nok"], "0.00%"),
    ]

    bg_kpi = [BLUE_LIGHT, BLUE_XLIGHT, BLUE_LIGHT, RED_LIGHT, RED_LIGHT]
    r = 3
    ws.cell(row=r, column=1, value="Indicateur").font  = hdr(bold=True, color=WHITE)
    ws.cell(row=r, column=1).fill = fill(BLUE_MID)
    ws.cell(row=r, column=1).alignment = center()
    ws.cell(row=r, column=2, value="Description").font = hdr(bold=True, color=WHITE)
    ws.cell(row=r, column=2).fill = fill(BLUE_MID)
    ws.cell(row=r, column=2).alignment = center()
    ws.cell(row=r, column=3, value="Valeur").font      = hdr(bold=True, color=WHITE)
    ws.cell(row=r, column=3).fill = fill(BLUE_MID)
    ws.cell(row=r, column=3).alignment = center()
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = border_thin()
    r += 1

    for i, (code, desc, val, fmt) in enumerate(kpis):
        ws.cell(row=r, column=1, value=code).font      = Font(name="Arial", bold=True, size=11)
        ws.cell(row=r, column=1).fill      = fill(bg_kpi[i])
        ws.cell(row=r, column=1).alignment = center()
        ws.cell(row=r, column=2, value=desc).font      = Font(name="Arial", size=10)
        ws.cell(row=r, column=2).fill      = fill(bg_kpi[i])
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=3, value=val).font       = Font(name="Arial", bold=True, size=11)
        ws.cell(row=r, column=3).fill      = fill(bg_kpi[i])
        ws.cell(row=r, column=3).alignment = center()
        if fmt:
            ws.cell(row=r, column=3).number_format = fmt
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = border_thin()
        r += 1

    # ── Données détail
    r += 1
    ws.cell(row=r, column=1, value="Détail des mouvements").font = hdr(bold=True, color=WHITE, size=11)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).fill      = fill(BLUE_MID)
    ws.cell(row=r, column=1).alignment = center()
    r += 1

    df_det = stats["df_detail"]
    det_cols = list(df_det.columns)

    for ci, col in enumerate(det_cols, 1):
        cell = ws.cell(row=r, column=ci, value=col)
        cell.font      = hdr(bold=True, color=WHITE, size=9)
        cell.fill      = fill(BLUE_DARK)
        cell.border    = border_thin()
        cell.alignment = center()
    r += 1

    for row_data in df_det.itertuples(index=False):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci)
            if isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
                cell.number_format = "YYYY/MM/DD HH:MM:SS"
            elif val is pd.NaT or (isinstance(val, float) and np.isnan(val)):
                cell.value = ""
            else:
                cell.value = val
            cell.font      = Font(name="Arial", size=9)
            cell.border    = border_thin()
            cell.alignment = Alignment(vertical="center")
            if det_cols[ci-1] in ("diff_day", "diff_off") and isinstance(cell.value, (int, float)):
                cell.fill = fill(GREEN_LIGHT if cell.value > 0 else (ORANGE_LIGHT if cell.value == 0 else RED_LIGHT))
        r += 1

    # Largeurs
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    for ci in range(1, len(det_cols)+1):
        col_letter = get_column_letter(ci)
        if ci > 3:
            ws.column_dimensions[col_letter].width = 22
    ws.freeze_panes = f"A{r - len(df_det)}"


def write_recap_global(wb, team_stats: dict):
    """Feuille récap globale toutes équipes."""
    ws = wb.create_sheet("Récap Global", 0)   # en premier
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"Récapitulatif Global — {TRIMESTRE}"
    c.font      = hdr(bold=True, color=WHITE, size=14)
    c.fill      = fill(BLUE_DARK)
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    headers = ["Équipe", "A — Somme diff_off", "B — Nb mouvements", "Moyenne (A/B)", "C — NOK", "% NOK (C/B)"]
    bg_h    = [BLUE_MID]*6
    r = 3
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font      = hdr(bold=True, color=WHITE, size=10)
        cell.fill      = fill(BLUE_MID)
        cell.border    = border_thin()
        cell.alignment = center()
    r += 1

    tot_A = tot_B = tot_C = 0
    for i, (team, s) in enumerate(sorted(team_stats.items())):
        bg = GRAY_LIGHT if i % 2 == 0 else WHITE
        vals = [team, s["A"], s["B"], s["moyenne"], s["C"], s["pct_nok"]]
        fmts = [None, None, None, "0.00", None, "0.00%"]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill(bg)
            cell.border    = border_thin()
            cell.alignment = Alignment(horizontal="left" if ci==1 else "center", vertical="center")
            if fmt:
                cell.number_format = fmt
            if ci in (5, 6) and isinstance(val, (int, float)) and val > 0:
                cell.fill = fill(RED_LIGHT)
        tot_A += s["A"]; tot_B += s["B"]; tot_C += s["C"]  # A=somme diff_off, B=nb lignes
        r += 1

    # Ligne TOTAL
    tot_moy  = round(tot_A / tot_B, 2) if tot_B else 0
    tot_pct  = round(tot_C / tot_B * 100, 2) if tot_B else 0
    totals   = ["TOTAL", tot_A, tot_B, tot_moy, tot_C, tot_pct]
    fmts_tot = [None, None, None, "0.00", None, "0.00%"]
    for ci, (val, fmt) in enumerate(zip(totals, fmts_tot), 1):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.font      = hdr(bold=True, color=WHITE, size=10)
        cell.fill      = fill(BLUE_DARK)
        cell.border    = border_thin()
        cell.alignment = Alignment(horizontal="left" if ci==1 else "center", vertical="center")
        if fmt:
            cell.number_format = fmt

    for col, w in [("A",25),("B",18),("C",16),("D",14),("E",12),("F",12)]:
        ws.column_dimensions[col].width = w

# ─────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────

# Styles
BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "BDD7EE"
BLUE_XLIGHT = "DEEAF1"
GREEN_LIGHT = "E2EFDA"
RED_LIGHT   = "FCE4D6"
ORANGE_LIGHT= "FFF2CC"
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F2F2F2"

def hdr(bold=True, color=WHITE, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")


def style_header_row(ws, row, start_col, end_col, bg=BLUE_MID, fg=WHITE, bold=True):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font   = hdr(bold=bold, color=fg)
        cell.fill   = fill(bg)
        cell.border = border_thin()
        cell.alignment = center()


def write_stat_block(ws, start_row, start_col, stats: dict, title_bg=BLUE_MID):
    r, c = start_row, start_col
    label = stats.get("label", "")

    # Titre
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
    cell = ws.cell(row=r, column=c, value=label)
    cell.font = hdr(bold=True, color=WHITE, size=12)
    cell.fill = fill(title_bg)
    cell.alignment = center()
    r += 1

    rows_data = [
        ("Nombre de lignes",  stats.get("count"),    None),
        ("Minimum (j)",       stats.get("min"),       None),
        ("Maximum (j)",       stats.get("max"),       None),
        ("Moyenne (j)",       stats.get("mean"),      "0.00"),
        ("Médiane (j)",       stats.get("median"),    "0.00"),
        ("90e percentile",    stats.get("p90"),       "0.00"),
        ("95e percentile",    stats.get("p95"),       "0.00"),
        ("Écart-type",        stats.get("std"),       "0.00"),
        ("Délai négatif",     stats.get("neg"),       None),
        ("Délai = 0",         stats.get("zero"),      None),
        ("Délai positif",     stats.get("pos"),       None),
    ]

    bg_map = {
        "Délai négatif": RED_LIGHT,
        "Délai = 0":     ORANGE_LIGHT,
        "Délai positif": GREEN_LIGHT,
    }

    for lbl, val, fmt in rows_data:
        bg_color = bg_map.get(lbl, GRAY_LIGHT if (r - start_row) % 2 == 0 else WHITE)
        lbl_cell = ws.cell(row=r, column=c, value=lbl)
        val_cell = ws.cell(row=r, column=c+1, value=val)
        for cell in (lbl_cell, val_cell):
            cell.font   = Font(name="Arial", size=10)
            cell.fill   = fill(bg_color)
            cell.border = border_thin()
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center")
        val_cell.alignment = right()
        if fmt:
            val_cell.number_format = fmt
        r += 1

    return r


def create_excel_report(df: pd.DataFrame, stats_day: dict, stats_off: dict, team_stats: dict, output_path: str):
    wb = Workbook()

    # ── Feuille 1 : Données ───────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Données"
    ws_data.freeze_panes = "A2"

    cols = list(df.columns)
    date_cols = {c for c in cols if df[c].dtype == "datetime64[ns]" or "date" in c.lower()}

    # En-têtes
    for ci, col in enumerate(cols, 1):
        cell = ws_data.cell(row=1, column=ci, value=col)
        cell.font   = hdr(bold=True, color=WHITE)
        cell.fill   = fill(BLUE_DARK)
        cell.border = border_thin()
        cell.alignment = center()

    # Données
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws_data.cell(row=ri, column=ci)
            if isinstance(val, pd.Timestamp):
                cell.value          = val.to_pydatetime()
                cell.number_format  = "YYYY/MM/DD HH:MM:SS"
            elif val is pd.NaT or (isinstance(val, float) and np.isnan(val)):
                cell.value = ""
            else:
                cell.value = val

            cell.border    = border_thin()
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")

            # Colorier diff_day / diff_off
            if cols[ci-1] in ("diff_day", "diff_off") and isinstance(cell.value, (int, float)):
                if cell.value < 0:
                    cell.fill = fill(RED_LIGHT)
                elif cell.value == 0:
                    cell.fill = fill(ORANGE_LIGHT)
                else:
                    cell.fill = fill(GREEN_LIGHT)

    # Largeurs colonnes
    for ci, col in enumerate(cols, 1):
        max_w = max(len(str(col)), 10)
        for ri in range(2, min(52, len(df)+2)):
            v = ws_data.cell(row=ri, column=ci).value
            if v:
                max_w = max(max_w, min(len(str(v)), 40))
        ws_data.column_dimensions[get_column_letter(ci)].width = max_w + 2

    # Filtre automatique
    ws_data.auto_filter.ref = ws_data.dimensions

    # ── Feuille 2 : Statistiques ──────────────────────────────────────────
    ws_stats = wb.create_sheet("Statistiques")
    ws_stats.sheet_view.showGridLines = False

    # Titre principal
    ws_stats.merge_cells("A1:F1")
    title_cell = ws_stats["A1"]
    title_cell.value     = f"Rapport HITS — {TRIMESTRE}"
    title_cell.font      = hdr(bold=True, color=WHITE, size=14)
    title_cell.fill      = fill(BLUE_DARK)
    title_cell.alignment = center()
    ws_stats.row_dimensions[1].height = 30

    # Infos
    info = [
        ("Trimestre",               TRIMESTRE),
        ("Lignes traitées",         stats_day.get("count", 0)),
        ("Jours fériés LU exclus",  11),
        ("Jours non fixes exclus",  len(JOURS_NON_FIXES)),
    ]
    r = 3
    for lbl, val in info:
        ws_stats.cell(row=r, column=1, value=lbl).font  = Font(name="Arial", bold=True, size=10)
        ws_stats.cell(row=r, column=2, value=val).font  = Font(name="Arial", size=10)
        ws_stats.cell(row=r, column=2).alignment        = right()
        r += 1

    # Blocs statistiques côte à côte
    r = write_stat_block(ws_stats, 8, 1, stats_day,  title_bg=BLUE_MID)
    write_stat_block(ws_stats,     8, 4, stats_off,  title_bg="1A6B3C")

    # Jours non fixes
    r2 = max(r, 8 + 13) + 2
    ws_stats.cell(row=r2, column=1, value="Jours non fixes configurés").font = hdr(bold=True, color=BLUE_DARK, size=10)
    r2 += 1
    if JOURS_NON_FIXES:
        for j in JOURS_NON_FIXES:
            ws_stats.cell(row=r2, column=1, value=j).font = Font(name="Arial", size=10)
            r2 += 1
    else:
        ws_stats.cell(row=r2, column=1, value="Aucun").font = Font(name="Arial", size=10, italic=True)

    # Largeurs
    for col, w in [("A", 28), ("B", 14), ("C", 4), ("D", 28), ("E", 14)]:
        ws_stats.column_dimensions[col].width = w

    # ── Feuille 3 : Graphiques ────────────────────────────────────────────
    ws_chart = wb.create_sheet("Graphiques")
    ws_chart.sheet_view.showGridLines = False

    # Données pour histogramme diff_off par tranche
    diff_off_vals = df["diff_off"].dropna()
    if len(diff_off_vals) > 0:
        vmin, vmax = int(diff_off_vals.min()), int(diff_off_vals.max())
        bins = list(range(vmin, vmax + 2, max(1, (vmax - vmin) // 10 + 1)))
        labels_b, counts_b = [], []
        for i in range(len(bins)-1):
            lo, hi = bins[i], bins[i+1]-1
            labels_b.append(f"{lo}-{hi}")
            counts_b.append(int(((diff_off_vals >= bins[i]) & (diff_off_vals < bins[i+1])).sum()))

        ws_chart["A1"] = "Tranche (diff_off)"
        ws_chart["B1"] = "Nombre de lignes"
        ws_chart["A1"].font = hdr(bold=True, color=WHITE)
        ws_chart["A1"].fill = fill(BLUE_MID)
        ws_chart["B1"].font = hdr(bold=True, color=WHITE)
        ws_chart["B1"].fill = fill(BLUE_MID)

        for i, (lbl, cnt) in enumerate(zip(labels_b, counts_b), 2):
            ws_chart.cell(row=i, column=1, value=lbl)
            ws_chart.cell(row=i, column=2, value=cnt)

        chart = BarChart()
        chart.type    = "col"
        chart.title   = "Distribution des délais (diff_off)"
        chart.y_axis.title = "Nombre de lignes"
        chart.x_axis.title = "Jours ouvrés"
        chart.style   = 10
        chart.width   = 20
        chart.height  = 12

        data_ref  = Reference(ws_chart, min_col=2, min_row=1, max_row=1+len(labels_b))
        cats_ref  = Reference(ws_chart, min_col=1, min_row=2, max_row=1+len(labels_b))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_chart.add_chart(chart, "D2")

    # ── Feuilles par équipe ──────────────────────────────────────────────
    if team_stats:
        write_recap_global(wb, team_stats)
        for team, s in sorted(team_stats.items()):
            write_team_sheet(wb, team, s)

    # ── Sauvegarde ────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n  Fichier Excel généré : {output_path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    print("=" * 55)
    print(f"  HITS — Traitement {TRIMESTRE}")
    print("=" * 55)

    print(f"\n[1] Chargement des fichiers depuis '{INPUT_FOLDER}/'")
    df_raw = load_and_fuse(INPUT_FOLDER)

    print("\n[2] Nettoyage et calcul des colonnes diff_day / diff_off")
    df, col_ud, col_ig = clean_and_compute(df_raw)

    print("\n[3] Calcul des statistiques")
    stats_day = compute_stats(df["diff_day"], "diff_day  (calendaire)")
    stats_off = compute_stats(df["diff_off"], "diff_off  (jours ouvrés LU)")

    print("\n  diff_day :")
    for k, v in stats_day.items():
        if k != "label":
            print(f"    {k:20s} {v}")

    print("\n  diff_off :")
    for k, v in stats_off.items():
        if k != "label":
            print(f"    {k:20s} {v}")

    print("\n[4] Statistiques par équipe (Team In-Charge)")
    nok_keys   = load_nok_keys(INPUT_FOLDER)
    team_stats = build_team_stats(df.copy(), nok_keys)
    for team, s in sorted(team_stats.items()):
        print(f"    {team:<25s}  A={s['A']}  B={s['B']}  Moy={s['moyenne']}  C={s['C']}  %NOK={s['pct_nok']}%")

    print("\n[5] Export Excel")
    output_path = os.path.join(OUTPUT_FOLDER, f"HITS_rapport_{TRIMESTRE}.xlsx")
    create_excel_report(df, stats_day, stats_off, team_stats, output_path)

    print("\n✓ Terminé.\n")


if __name__ == "__main__":
    main()
